import os
import io
import re
import base64
from datetime import datetime, timezone
from typing import Optional, Dict, Any, List, Tuple

import requests
import pandas as pd
from msal import ConfidentialClientApplication
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


# =========================================================
# CONFIG
# =========================================================
GRAPH = "https://graph.microsoft.com/v1.0"

SUBJECT_PHRASE = "Weekly Time Utilization"
MAILBOX_UPN = os.getenv("MAILBOX_UPN", "apatil@glassdoctordfw.com")
SEND_TO = "8994d80f.glassdoctordfw.com@amer.teams.ms"

INPUT_SHEET_NAME = "Sheet1"

# Rename these BU values to Retail in BOTH column R and S
RETAIL_MAP_VALUES = {"Arlington", "Colleyville", "Carrollton", "Dallas", "Denton"}
RETAIL_REPLACEMENT = "Retail"

# Excluded employee names
EXCLUDED_EMPLOYEES = {
    "Andres Segura",
    "Edwin Segura",
    "Javier David Colorado Perez",
    "Jorge Diaz",
    "Leandro Espinosa",
    "Tommy Bartholomew",
}

OUTPUT_FILENAME = "Weekly Time Summary.xlsx"


# =========================================================
# COLOR FILLS
# =========================================================
SUPER_GREEN_FILL = PatternFill(fill_type="solid", fgColor="00B050")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="92D050")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
RED_FILL = PatternFill(fill_type="solid", fgColor="FF0000")


# =========================================================
# HELPERS
# =========================================================
def parse_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)
    if dt_str.endswith("Z"):
        dt_str = dt_str.replace("Z", "+00:00")
    return datetime.fromisoformat(dt_str)


def get_token() -> str:
    tenant_id = os.getenv("tenant_id")
    client_id = os.getenv("client_id")
    client_secret = os.getenv("client_secret")
    if not tenant_id:
        raise RuntimeError("Missing environment variable: tenant_id")
    if not client_id:
        raise RuntimeError("Missing environment variable: client_id")
    if not client_secret:
        raise RuntimeError("Missing environment variable: client_secret")

    app = ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
    )

    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error')} - {result.get('error_description')}")
    return result["access_token"]


def graph_get(token: str, url: str, params: Optional[dict] = None) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }
    r = requests.get(url, headers=headers, params=params, timeout=60)
    if not r.ok:
        try:
            print("Graph GET error payload:", r.json())
        except Exception:
            print("Graph GET error text:", r.text)
        r.raise_for_status()
    return r.json()


def graph_get_bytes(token: str, url: str, params: Optional[dict] = None) -> bytes:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }
    r = requests.get(url, headers=headers, params=params, timeout=120)
    if not r.ok:
        try:
            print("Graph GET bytes error payload:", r.json())
        except Exception:
            print("Graph GET bytes error text:", r.text)
        r.raise_for_status()
    return r.content


def graph_post(token: str, url: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    r = requests.post(url, headers=headers, json=payload, timeout=120)
    if not r.ok:
        try:
            print("Graph POST error payload:", r.json())
        except Exception:
            print("Graph POST error text:", r.text)
        r.raise_for_status()
    if r.text.strip():
        return r.json()
    return {}


def latest_message_for_subject(token: str, mailbox_upn: str, subject_phrase: str) -> Optional[Dict[str, Any]]:
    """
    Finds the latest message whose subject contains the phrase.
    """
    url = f"{GRAPH}/users/{mailbox_upn}/mailFolders/Inbox/messages"
    params = {
        "$select": "id,subject,receivedDateTime,from,hasAttachments",
        "$top": "25",
        "$search": f"\"{subject_phrase}\"",
    }

    data = graph_get(token, url, params=params)
    msgs: List[Dict[str, Any]] = data.get("value", [])

    phrase = subject_phrase.lower()
    candidates = [m for m in msgs if phrase in (m.get("subject") or "").lower()]

    if not candidates:
        return None

    candidates.sort(key=lambda m: parse_dt(m.get("receivedDateTime", "")), reverse=True)
    return candidates[0]


def get_first_xlsx_attachment_from_message(
    token: str, mailbox_upn: str, message_id: str
) -> Tuple[Optional[str], Optional[bytes]]:
    """
    Returns the first .xlsx attachment from the email.
    """
    url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments"
    data = graph_get(token, url, params={"$top": "50"})
    atts = data.get("value", [])

    for a in atts:
        name = a.get("name") or ""
        if name.lower().endswith(".xlsx"):
            content_bytes = a.get("contentBytes")
            if content_bytes:
                return name, base64.b64decode(content_bytes)

            att_id = a.get("id")
            if att_id:
                raw_url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments/{att_id}/$value"
                return name, graph_get_bytes(token, raw_url)

    return None, None


def normalize_bu_value(val):
    if pd.isna(val):
        return val
    text = str(val).strip()
    if text in RETAIL_MAP_VALUES:
        return RETAIL_REPLACEMENT
    return text


def autosize_worksheet(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 35)


def get_fill_for_metric(metric_name: str, value: float):
    """
    Returns the correct fill based on the metric and percentage thresholds.
    """
    if value is None:
        return None

    try:
        value = float(value)
    except Exception:
        return None

    metric_name = metric_name.strip().lower()

    if metric_name == "% working hours":
        if value > 70:
            return SUPER_GREEN_FILL
        elif 60 <= value <= 70:
            return GREEN_FILL
        elif 50 <= value < 60:
            return YELLOW_FILL
        else:
            return RED_FILL

    elif metric_name == "% idle hours":
        if value < 10:
            return SUPER_GREEN_FILL
        elif 10 <= value < 15:
            return GREEN_FILL
        elif 15 <= value <= 20:
            return YELLOW_FILL
        else:
            return RED_FILL

    elif metric_name == "% driving hours":
        if value < 10:
            return SUPER_GREEN_FILL
        elif 10 <= value < 15:
            return GREEN_FILL
        elif 15 <= value <= 25:
            return YELLOW_FILL
        else:
            return RED_FILL

    return None


def apply_conditional_colors(ws):
    """
    Applies fill colors to % Working Hours, % Idle Hours, % Driving Hours columns.
    """
    header_map = {}
    for cell in ws[1]:
        header_map[cell.value] = cell.column

    target_headers = ["% Working Hours", "% Idle Hours", "% Driving Hours"]

    for header in target_headers:
        col_idx = header_map.get(header)
        if not col_idx:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            fill = get_fill_for_metric(header, cell.value)
            if fill:
                cell.fill = fill


def send_email_with_attachments(
    token: str,
    sender_upn: str,
    to_email: str,
    subject: str,
    body_text: str,
    attachments: List[Tuple[str, bytes]],
):
    graph_attachments = []
    for filename, file_bytes in attachments:
        graph_attachments.append({
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": filename,
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "contentBytes": base64.b64encode(file_bytes).decode("utf-8"),
        })

    payload = {
        "message": {
            "subject": subject,
            "body": {
                "contentType": "Text",
                "content": body_text,
            },
            "toRecipients": [
                {"emailAddress": {"address": to_email}}
            ],
            "attachments": graph_attachments,
        },
        "saveToSentItems": True,
    }

    url = f"{GRAPH}/users/{sender_upn}/sendMail"
    graph_post(token, url, payload)


# =========================================================
# CORE LOGIC
# =========================================================
def build_summary_file(input_excel_bytes: bytes) -> Tuple[bytes, str]:
    """
    Reads input Excel, normalizes BU columns, excludes specific employees,
    builds summary workbook with percentage split only, returns output bytes + filename.
    """
    df = pd.read_excel(io.BytesIO(input_excel_bytes), sheet_name=INPUT_SHEET_NAME)

    required_columns = [
        "Employee Name",
        "Activity",
        "Reg Hours",
        "OT Hours",
        "Business Unit",
        "Employee Business Unit",
    ]
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        raise RuntimeError(f"Missing expected columns in Sheet1: {missing}")

    # Clean names first
    df["Employee Name"] = df["Employee Name"].astype(str).str.strip()

    # Exclude employees
    df = df[~df["Employee Name"].isin(EXCLUDED_EMPLOYEES)].copy()

    # Normalize BU values in both columns R and S
    df["Business Unit"] = df["Business Unit"].apply(normalize_bu_value)
    df["Employee Business Unit"] = df["Employee Business Unit"].apply(normalize_bu_value)

    # Standardize activity
    df["Activity"] = df["Activity"].astype(str).str.strip().str.title()

    # Keep only Idle / Working / Driving
    df = df[df["Activity"].isin(["Idle", "Working", "Driving"])].copy()

    # Hours = Reg Hours + OT Hours
    df["Reg Hours"] = pd.to_numeric(df["Reg Hours"], errors="coerce").fillna(0)
    df["OT Hours"] = pd.to_numeric(df["OT Hours"], errors="coerce").fillna(0)
    df["Total Hours"] = df["Reg Hours"] + df["OT Hours"]

    # -----------------------------------------------------
    # Employee Summary
    # -----------------------------------------------------
    emp_pivot = (
        df.pivot_table(
            index="Employee Name",
            columns="Activity",
            values="Total Hours",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    for col in ["Idle", "Working", "Driving"]:
        if col not in emp_pivot.columns:
            emp_pivot[col] = 0.0

    emp_pivot = emp_pivot[["Employee Name", "Idle", "Working", "Driving"]]

    emp_pivot["Total"] = emp_pivot["Idle"] + emp_pivot["Working"] + emp_pivot["Driving"]

    emp_pivot["% Idle Hours"] = (
        emp_pivot["Idle"] / emp_pivot["Total"].replace(0, pd.NA) * 100
    ).fillna(0).round(2)

    emp_pivot["% Working Hours"] = (
        emp_pivot["Working"] / emp_pivot["Total"].replace(0, pd.NA) * 100
    ).fillna(0).round(2)

    emp_pivot["% Driving Hours"] = (
        emp_pivot["Driving"] / emp_pivot["Total"].replace(0, pd.NA) * 100
    ).fillna(0).round(2)

    emp_summary = emp_pivot[[
        "Employee Name",
        "% Idle Hours",
        "% Working Hours",
        "% Driving Hours"
    ]].sort_values("Employee Name").reset_index(drop=True)

    # -----------------------------------------------------
    # Business Unit Summary
    # -----------------------------------------------------
    bu_employee_activity = (
        df.groupby(["Business Unit", "Employee Name", "Activity"], dropna=False)["Total Hours"]
        .sum()
        .reset_index()
    )

    bu_avg = (
        bu_employee_activity.pivot_table(
            index=["Business Unit", "Employee Name"],
            columns="Activity",
            values="Total Hours",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    for col in ["Idle", "Working", "Driving"]:
        if col not in bu_avg.columns:
            bu_avg[col] = 0.0

    bu_summary_raw = (
        bu_avg.groupby("Business Unit", dropna=False)[["Idle", "Working", "Driving"]]
        .mean()
        .reset_index()
    )

    bu_summary_raw["Total"] = (
        bu_summary_raw["Idle"] +
        bu_summary_raw["Working"] +
        bu_summary_raw["Driving"]
    )

    bu_summary_raw["% Idle Hours"] = (
        bu_summary_raw["Idle"] / bu_summary_raw["Total"].replace(0, pd.NA) * 100
    ).fillna(0).round(2)

    bu_summary_raw["% Working Hours"] = (
        bu_summary_raw["Working"] / bu_summary_raw["Total"].replace(0, pd.NA) * 100
    ).fillna(0).round(2)

    bu_summary_raw["% Driving Hours"] = (
        bu_summary_raw["Driving"] / bu_summary_raw["Total"].replace(0, pd.NA) * 100
    ).fillna(0).round(2)

    bu_summary = bu_summary_raw[[
        "Business Unit",
        "% Idle Hours",
        "% Working Hours",
        "% Driving Hours"
    ]].sort_values("Business Unit").reset_index(drop=True)

    # -----------------------------------------------------
    # Write output workbook
    # -----------------------------------------------------
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        bu_summary.to_excel(writer, sheet_name="Business Unit Summary", index=False)
        emp_summary.to_excel(writer, sheet_name="Employee Summary", index=False)

        wb = writer.book
        ws1 = wb["Business Unit Summary"]
        ws2 = wb["Employee Summary"]

        apply_conditional_colors(ws1)
        apply_conditional_colors(ws2)

        autosize_worksheet(ws1)
        autosize_worksheet(ws2)

    return output.getvalue(), OUTPUT_FILENAME


# =========================================================
# MAIN
# =========================================================
def main():
    token = get_token()

    print(f"Searching latest email with subject containing: {SUBJECT_PHRASE}")
    msg = latest_message_for_subject(token, MAILBOX_UPN, SUBJECT_PHRASE)
    if not msg:
        print("No email found with that subject.")
        return

    subject = msg.get("subject", SUBJECT_PHRASE)
    message_id = msg["id"]

    print(f"Found email: {subject}")

    input_filename, input_bytes = get_first_xlsx_attachment_from_message(token, MAILBOX_UPN, message_id)
    if not input_bytes:
        print("No .xlsx attachment found on the email.")
        return

    print(f"Downloaded attachment: {input_filename}")

    summary_bytes, summary_filename = build_summary_file(input_bytes)

    # Save locally
    with open(input_filename, "wb") as f:
        f.write(input_bytes)

    with open(summary_filename, "wb") as f:
        f.write(summary_bytes)

    print(f"Created summary file: {summary_filename}")

    body_text = (
        "Hi,\n\n"
        "Please find attached the Weekly Time Summary file.\n\n"
        "The summary workbook includes:\n"
        "- Business Unit Summary\n"
        "- Employee Summary\n\n"
        "Thanks,\n"
        "Aayush"
    )

    send_email_with_attachments(
        token=token,
        sender_upn=MAILBOX_UPN,
        to_email=SEND_TO,
        subject=subject,
        body_text=body_text,
        attachments=[
            (summary_filename, summary_bytes),
        ],
    )

    print("Email sent successfully.")


if __name__ == "__main__":
    main()
